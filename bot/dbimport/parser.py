"""Разбор файла БД на записи: SQLite, CSV, JSON (в т.ч. Telegram export), XLSX,
а также архивы (ZIP/RAR), текстовые файлы и .torrent раздачи."""
import csv
import json
import logging
import os
import re
import shutil
import sqlite3
import subprocess
import tempfile
import zipfile
from pathlib import Path

from .schema import (
    detect_vertical_matrix,
    has_contact_columns,
    horizontal_records,
    split_txt_records,
    summarize_contacts,
    vertical_matrix_records,
)
from .torrent import parse_torrent

logger = logging.getLogger(__name__)

# Расширения, которые бот принимает в /import
SUPPORTED_EXTENSIONS = {".db", ".sqlite", ".sqlite3", ".sql", ".csv", ".json",
                        ".xlsx", ".zip", ".rar", ".7z", ".torrent", ".txt"}

# Ограничения при распаковке архивов (защита от «архивных» атак).
# Настраиваются переменными окружения (Render → Environment):
#   MAX_ARCHIVE_FILES   — максимум файлов в архиве (по умолчанию 1000)
#   MAX_ARCHIVE_TOTAL_MB — суммарный объём распакованных данных (по умолчанию 1024)
MAX_ARCHIVE_FILES = int(os.getenv("MAX_ARCHIVE_FILES", "1000"))
MAX_ARCHIVE_TOTAL_MB = int(os.getenv("MAX_ARCHIVE_TOTAL_MB", "1024"))
MAX_ARCHIVE_TOTAL = MAX_ARCHIVE_TOTAL_MB * 1024 * 1024
MAX_ARCHIVE_DEPTH = 3

# Колонки, которые ищем в данных (основные русские и английские варианты)
_TEXT_KEYS = ("text", "message", "content", "title", "description", "body",
              "comment", "tresc", "post")
_AUTHOR_KEYS = ("author", "from", "sender", "user", "username", "channel",
                "channel_name", "nickname", "nick", "автор")
_DATE_KEYS = ("date", "datetime", "created_at", "published_at", "time", "ts",
              "date_time", "дата")
_URL_KEYS = ("url", "link", "source_url", "permalink", "href")

# Колонка с явной категорией/субъектом — её значение становится разделом
_CATEGORY_KEYS = ("категория", "категори", "раздел", "рубрика", "тема", "темы",
                  "группа", "группы", "тип", "направление", "отрасль", "вид",
                  "субъект", "субъекты", "category", "categories", "subject",
                  "subjects", "section", "sections", "group", "type", "topic")

# Заголовки, которые не подходят как название раздела
_GENERIC_HEADERS = {h.lower() for h in (
    "text", "message", "content", "title", "description", "body", "comment",
    "post", "tresc", "сообщение", "текст", "содержание", "описание",
    "заголовок", "комментарий")}

_META_HEADERS = ("no", "n", "id", "номер", "код", "индекс", "порядковый",
                 "number", "index")
_SKIP_HEADERS = {h.lower() for h in
                 (*_DATE_KEYS, *_URL_KEYS, *_AUTHOR_KEYS, *_META_HEADERS)} | {"№", "#"}

MAX_TEXT = 20000


def parse_file(path: str | Path, filename: str = "", _depth: int = 0) -> tuple[dict, list[dict]]:
    """Парсит файл и возвращает (метаинформация, записи).

    Каждая запись — dict: {source, author, text, url, date}.
    """
    path = Path(path)
    ext = path.suffix.lower()
    if ext in (".zip", ".rar", ".7z"):
        return _parse_archive(path, filename or path.name, _depth)
    if ext == ".torrent":
        parsed = parse_torrent(path, filename or path.name)
        return parsed["meta"], parsed["records"]
    if ext in (".db", ".sqlite", ".sqlite3"):
        rows, source = _iter_sqlite(path)
        fmt = "sqlite"
    elif ext == ".sql":
        rows, source = _iter_sql_dump(path)
        fmt = "sql"
    elif ext == ".csv":
        rows, source = _iter_csv(path)
        fmt = "csv"
    elif ext == ".json":
        rows, source = _iter_json(path)
        fmt = "json"
    elif ext == ".xlsx":
        rows, source = _iter_xlsx(path)
        fmt = "xlsx"
    elif ext == ".txt":
        rows, source = _iter_txt(path)
        fmt = "txt"
    else:
        raise ValueError(f"Неподдерживаемый формат: {ext}. Поддерживаются: "
                         f"{', '.join(sorted(SUPPORTED_EXTENSIONS))}")

    records: list[dict] = []

    if fmt == "txt":
        lines = [str(r.get("text", "")) for r in rows]
        smart = split_txt_records(lines)
        if smart is not None:
            records = smart
        else:
            records = [{"text": ln} for ln in lines if ln.strip()]
    elif rows and isinstance(rows[0], dict) and has_contact_columns(rows[0].keys()):
        records = horizontal_records(rows, source)
    elif rows and isinstance(rows[0], dict) and detect_vertical_matrix(rows):
        records = vertical_matrix_records(rows, source)

    if not records and rows and isinstance(rows[0], dict):
        columns_mode = _columns_mode(rows)
        for row in rows:
            if columns_mode:
                recs = _normalize_columns(row, source)
            else:
                rec = _normalize(row, source)
                recs = [rec] if rec else []
            for rec in recs:
                rec["source"] = rec["source"] or (filename or path.name)
                if fmt == "txt" and not rec.get("section"):
                    rec["section"] = (filename or path.stem)[:255]
                records.append(rec)

    for rec in records:
        rec["source"] = rec.get("source") or (filename or path.name)
        if fmt == "txt" and not rec.get("section"):
            rec["section"] = (filename or path.stem)[:255]

    meta = {"format": fmt, "source": source, "filename": filename or path.name,
            "rows_parsed": len(rows), "rows_kept": len(records)}
    meta["contacts"] = summarize_contacts(records)
    logger.info("parse %s: %s -> %s записей (contacts=%s)",
                filename, fmt, len(records), meta["contacts"])
    return meta, records


# ---------- архивы ----------

def _safe_member(name: str) -> str:
    """Имя члена архива без «побегов» наружу (zip-slip)."""
    parts = [p for p in (name or "").replace("\\", "/").split("/")
             if p and p not in (".", "..") and not p.startswith(":")]
    return "/".join(parts)


def _extract_zip(path: Path, target: Path):
    total = 0
    with zipfile.ZipFile(str(path)) as archive:
        members = [i for i in archive.infolist() if not i.is_dir()]
        if len(members) > MAX_ARCHIVE_FILES:
            raise ValueError(f"В архиве больше {MAX_ARCHIVE_FILES} файлов")
        for info in members:
            total += info.file_size
            if total > MAX_ARCHIVE_TOTAL:
                raise ValueError(f"Архив больше {MAX_ARCHIVE_TOTAL_MB} МБ — не разбираю")
            fn = _safe_member(info.filename)
            if not fn:
                continue
            dest = target / fn
            dest.parent.mkdir(parents=True, exist_ok=True)
            with archive.open(info) as src, open(dest, "wb") as out:
                shutil.copyfileobj(src, out, 1024 * 256)


def _extract_rar(path: Path, target: Path):
    # 1) rarfile + системный unrar/bsdtar
    try:
        import rarfile
        with rarfile.RarFile(str(path)) as archive:
            infos = [i for i in archive.infolist() if not i.isdir()]
            if len(infos) > MAX_ARCHIVE_FILES:
                raise ValueError(f"В архиве больше {MAX_ARCHIVE_FILES} файлов")
            total = sum(i.file_size for i in infos)
            if total > MAX_ARCHIVE_TOTAL:
                raise ValueError(f"Архив больше {MAX_ARCHIVE_TOTAL_MB} МБ — не разбираю")
            for info in infos:
                fn = _safe_member(info.filename)
                if not fn:
                    continue
                dest = target / fn
                dest.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(info) as src, open(dest, "wb") as out:
                    shutil.copyfileobj(src, out, 1024 * 256)
        return
    except ValueError:
        raise
    except Exception as ex:
        logger.info("rarfile: %s — пробую bsdtar", ex)

    # 2) запасной путь: bsdtar (libarchive)
    try:
        proc = subprocess.run(["bsdtar", "-xf", str(path), "-C", str(target)],
                              capture_output=True, timeout=120)
    except FileNotFoundError:
        raise ValueError(
            "RAR: не найден распаковщик (bsdtar/unrar не установлен)") from None
    if proc.returncode != 0:
        raise ValueError(
            "Не удалось распаковать RAR: возможно, файл повреждён или с паролем "
            "(в контейнере: apt-get install libarchive-tools)")


def _extract_7z(path: Path, target: Path):
    # 1) py7zr — чистый Python, работает и в Windows, и в контейнере
    try:
        import py7zr
        with py7zr.SevenZipFile(str(path), mode="r") as archive:
            names = archive.getnames()
            files = [n for n in names if not n.endswith("/")]
            if len(files) > MAX_ARCHIVE_FILES:
                raise ValueError(f"В архиве больше {MAX_ARCHIVE_FILES} файлов")
            # Суммарный размер через list()
            total = 0
            for info in archive.list():
                if info.is_directory:
                    continue
                total += info.uncompressed or 0
            if total > MAX_ARCHIVE_TOTAL:
                raise ValueError(f"Архив больше {MAX_ARCHIVE_TOTAL_MB} МБ — не разбираю")
            archive.extractall(path=target)
        return
    except ValueError:
        raise
    except Exception as ex:
        logger.info("py7zr: %s — пробую 7z/bsdtar", ex)

    # 2) запасной путь: 7z / 7za / 7zr / bsdtar
    for tool in ("7z", "7za", "7zr", "bsdtar"):
        if tool == "bsdtar":
            cmd = [tool, "-xf", str(path), "-C", str(target)]
        else:
            cmd = [tool, "x", "-y", f"-o{target}", str(path)]
        try:
            proc = subprocess.run(cmd, capture_output=True, timeout=180)
        except FileNotFoundError:
            continue
        if proc.returncode == 0:
            return
        # 7z вернул ошибку — пробуем следующий
    raise ValueError(
        "7z: не найден распаковщик (py7zr не установлен и нет 7z/bsdtar)")


def _parse_archive(path: Path, filename: str, depth: int) -> tuple[dict, list[dict]]:
    if depth >= MAX_ARCHIVE_DEPTH:
        raise ValueError("Вложенность архивов больше 3 уровней")
    tmp = Path(tempfile.mkdtemp(prefix="import_"))
    try:
        ext = path.suffix.lower()
        if ext == ".zip":
            _extract_zip(path, tmp)
        elif ext == ".7z":
            _extract_7z(path, tmp)
        else:
            _extract_rar(path, tmp)

        records: list[dict] = []
        inner = [p for p in tmp.rglob("*") if p.is_file()]
        for f in inner:
            try:
                _, recs = parse_file(f, f.name, _depth=depth + 1)
            except (ValueError, OSError) as ex:
                logger.info("внутренний файл %s пропущен: %s", f.name, ex)
                continue
            records.extend(recs)

        meta = {"format": path.suffix.lower().lstrip("."), "source": path.name,
                "filename": filename, "rows_parsed": len(inner),
                "rows_kept": len(records)}
        return meta, records
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def _row_keymap(row: dict) -> dict[str, str]:
    """Заголовок(нижний регистр) → оригинальное имя колонки."""
    mapping: dict[str, str] = {}
    for key in row:
        norm = str(key).strip()
        low = norm.lower()
        if low and low not in mapping:
            mapping[low] = norm
    return mapping


def _find_ci(row: dict, keys) -> object:
    """Значение колонки по имени без учёта регистра."""
    mapping = _row_keymap(row)
    for key in keys:
        orig = mapping.get(key.lower())
        if orig is not None:
            return row[orig]
    return None


def _columns_mode(rows: list[dict]) -> bool:
    """True, если колонки не «семантические» — тогда каждая колонка = раздел."""
    if not rows:
        return False
    semantic = {h.lower() for h in
                (*_TEXT_KEYS, *_CATEGORY_KEYS, *_AUTHOR_KEYS, *_DATE_KEYS, *_URL_KEYS)}
    keys = {str(k).strip().lower() for k in rows[0].keys()}
    return not bool(keys & semantic)


def _normalize_columns(row: dict, default_source: str) -> list[dict]:
    """Одна запись на колонку: раздел = название столбца."""
    out: list[dict] = []
    for key, val in row.items():
        if key in ("", None):
            continue
        header = str(key).strip()
        norm = header.lower()
        if norm.startswith("col") and norm[3:].isdigit():
            continue
        if norm in _SKIP_HEADERS or norm in _GENERIC_HEADERS:
            continue
        text = _join_parts(val)
        if not text or text.isdigit():
            continue
        out.append({
            "source": str(default_source or ""),
            "section": header[:255],
            "author": "",
            "text": text[:MAX_TEXT],
            "url": "",
            "date": "",
        })
    return out


def _clean(text) -> str:
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    return " ".join(text.split())


def _parse_date(value) -> str:
    """Нормализуем дату в строку ISO (без жёсткого парсинга)."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    return s[:50]


def _join_parts(val) -> str:
    """Текст из списка (Telegram export: list из строк и dict)."""
    if isinstance(val, list):
        parts = []
        for part in val:
            if isinstance(part, dict):
                parts.append(str(part.get("text", part.get("message", "")) or ""))
            else:
                parts.append(str(part))
        return " ".join(p for p in parts if p and p.strip())
    return _clean(val)


def _row_text_section(row: dict) -> tuple[str, str]:
    """(текст, раздел) записи.

    Раздел берётся приоритетно из явной колонки категории/субъекта.
    Если её нет, а текст лежит в колонке с осмысленным заголовком —
    разделом становится НАЗВАНИЕ СТОЛБЦА, где лежит текст.
    """
    section = ""
    cat_val = _find_ci(row, _CATEGORY_KEYS)
    if cat_val is not None and str(cat_val).strip():
        section = _clean(str(cat_val))[:255]

    # 1) известная текстовая колонка
    keymap = _row_keymap(row)
    for key in _TEXT_KEYS:
        orig = keymap.get(key.lower())
        if orig is None:
            continue
        val = _join_parts(row[orig])
        if val:
            return val, section

    # 2) fallback: любой непустой столбец, заголовок → раздел
    best_key, best = None, None
    for key, val in row.items():
        norm_key = str(key).strip().lower()
        if key in ("", None) or (norm_key.startswith("col") and norm_key[3:].isdigit()):
            continue
        if norm_key in _SKIP_HEADERS:
            continue
        sval = _join_parts(val)
        if not sval or sval.isdigit():
            continue
        if best is None or len(sval) > len(best):
            best, best_key = sval, key

    if best:
        if not section and str(best_key).strip().lower() not in _GENERIC_HEADERS:
            section = str(best_key).strip()[:255]
        return best, section
    return "", section


def _normalize(row: dict, default_source: str) -> dict:
    text, section = _row_text_section(row)
    if not text:
        return {}

    author = _find_ci(row, _AUTHOR_KEYS)
    if isinstance(author, dict):  # Telegram export: from = {"id":..., "name":...}
        author = author.get("name") or author.get("username") or ""
    url = _find_ci(row, _URL_KEYS)
    if isinstance(url, list):
        url = ", ".join(str(u) for u in url if u)

    return {
        "source": str(default_source or ""),
        "section": section or "",
        "author": _clean(str(author or "")),
        "text": text[:MAX_TEXT],
        "url": _clean(str(url or "")),
        "date": _parse_date(_find_ci(row, _DATE_KEYS)),
    }


def _iter_sqlite(path: Path) -> tuple[list[dict], str]:
    """Читаем все таблицы SQLite, отдаём строки, где есть текст."""
    rows: list[dict] = []
    source = ""
    def _q(name: str) -> str:
        return '"' + str(name).replace('"', '""') + '"'

    try:
        conn = sqlite3.connect(str(path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        tables = [row[0] for row in cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table'")]
        for table in tables:
            try:
                cols = [c[1] for c in cur.execute(
                    f'PRAGMA table_info({_q(table)})')]
            except sqlite3.Error:
                continue
            text_cols = [c for c in cols if c.lower() in _TEXT_KEYS]
            select = ", ".join(_q(c) for c in cols)
            for row in cur.execute(f'SELECT {select} FROM {_q(table)}'):
                d = dict(zip(cols, row))
                if text_cols and any(d.get(c) for c in text_cols):
                    rows.append(d)
                    if not source:
                        source = table
            # Если таблица пустая — просто не попадает в записи
        cur.close()
        conn.close()
    except Exception as ex:
        logger.warning("sqlite parse error: %s", ex)
    if not source:
        source = "sqlite"
    return rows, source


_SQL_TOKEN_VALUES = re.compile(
    r"INSERT\s+INTO\s+`?([\wа-яёА-ЯЁ\-]+)`?\s*\(([^)]+)\)\s*VALUES",
    re.IGNORECASE,
)
_SQL_SIMPLE_INSERT = re.compile(
    r"INSERT\s+INTO\s+`?([\wа-яёА-ЯЁ\-]+)`?(?:\s*\(([^)]*)\))?\s*VALUES",
    re.IGNORECASE,
)


def _sql_split_values(s: str) -> list[str]:
    """Разбивает список значений INSERT (a,b,(c,d),(e,f)) на отдельные кортежи."""
    values: list[str] = []
    i = 0
    depth = 0
    start = 0
    in_str = None
    while i < len(s):
        ch = s[i]
        if in_str:
            if ch == "\\":
                i += 2
                continue
            if ch == in_str:
                in_str = None
            i += 1
            continue
        if ch in "'\"":
            in_str = ch
            i += 1
            continue
        if ch == "(":
            if depth == 0:
                start = i
            depth += 1
        elif ch == ")":
            depth -= 1
            if depth == 0:
                values.append(s[start:i + 1])
        elif ch == "," and depth == 0:
            pass
        i += 1
    return values


def _sql_parse_value(tok: str):
    """Распарсить один SQL литерал в Python-значение."""
    tok = tok.strip()
    if not tok:
        return ""
    single = False
    if len(tok) >= 2 and tok[0] == tok[-1] and tok[0] in "'\"":
        single = True
        body = tok[1:-1].replace("''", "'").replace('\\"', '"')
        return body
    low = tok.lower()
    if low in ("null", "none", ""):
        return ""
    if low in ("true", "1", "b'1'"):
        return 1
    if low in ("false", "0", "b'0'"):
        return 0
    if single:
        return ""
    try:
        return int(tok)
    except ValueError:
        pass
    try:
        return float(tok)
    except ValueError:
        pass
    return tok


def _parse_sql_row(values_str: str, cols: list[str] | None,
                   table: str) -> dict:
    """Собирает dict: колонка -> значение из кортежа VALUES."""
    inner = values_str[1:-1].strip()
    # Простейшее разделение на токены с учётом строк в кавычках
    tokens: list[str] = []
    cur = ""
    in_str = None
    for ch in inner:
        if in_str:
            cur += ch
            if ch == "\\":
                continue
            if ch == in_str:
                in_str = None
            continue
        if ch in "'\"":
            in_str = ch
            cur += ch
        elif ch == ",":
            tokens.append(cur)
            cur = ""
        else:
            cur += ch
    if cur.strip() or tokens:
        tokens.append(cur)

    if cols:
        d = {}
        for i, tok in enumerate(tokens):
            col = cols[i] if i < len(cols) else f"col{i + 1}"
            d[col] = _sql_parse_value(tok)
        return d
    # Без заголовков колонок — нумеруем
    return {f"col{i + 1}": _sql_parse_value(tok) for i, tok in enumerate(tokens)}


def _iter_sql_dump(path: Path) -> tuple[list[dict], str]:
    """Читает SQL-дамп: таблицы и строки INSERT ... VALUES."""
    rows: list[dict] = []
    source = ""
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            text = f.read()
    except OSError as ex:
        logger.warning("sql dump read error: %s", ex)
        return [], "sql"

    for m in _SQL_SIMPLE_INSERT.finditer(text):
        table = m.group(1)
        cols_raw = m.group(2) or ""
        cols = [c.strip().strip("`\"'") for c in cols_raw.split(",") if c.strip()] or None
        values_str = text[m.end():]
        # отрезаем до конца VALUES-списка (до ';' на верхнем уровне)
        end = 0
        depth = 0
        in_str = None
        for i in range(len(values_str)):
            ch = values_str[i]
            if in_str:
                if ch == "\\":
                    i += 1
                    continue
                if ch == in_str:
                    in_str = None
                continue
            if ch in "'\"":
                in_str = ch
            elif ch == "(":
                depth += 1
            elif ch == ")":
                depth -= 1
            elif ch == ";" and depth <= 0:
                end = i
                break
        if end == 0:
            end = len(values_str)
        segment = values_str[:end]
        # Собираем кортежи VALUES (....), (....), ...
        for tuple_str in _sql_split_values(segment):
            try:
                row = _parse_sql_row(tuple_str, cols, table)
            except Exception as ex:
                logger.debug("sql row parse: %s", ex)
                continue
            if any((v is not None and str(v).strip()) for v in row.values()):
                rows.append(row)
                if not source:
                    source = table
    if not source:
        source = "sql"
    return rows, source


def _iter_csv(path: Path) -> tuple[list[dict], str]:
    rows: list[dict] = []
    source = "csv"
    with open(path, "r", encoding="utf-8-sig", errors="replace", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        for row in csv.DictReader(f, dialect=dialect):
            if row and any((v or "").strip() for v in row.values()):
                rows.append(row)
    return rows, source


def _iter_json(path: Path) -> tuple[list[dict], str]:
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        data = json.load(f)

    items: list = []
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        # Telegram export → ключ "messages"
        for key in ("messages", "items", "posts", "rows", "data", "result"):
            if isinstance(data.get(key), list):
                items = data[key]
                break
        if not items:
            items = [data]
    source = "json"
    if isinstance(data, dict) and data.get("chat"):
        source = _clean(str(data["chat"].get("title", "") or "")).strip() or "json"

    rows = [item for item in items if isinstance(item, dict)]
    return rows, source


def _iter_xlsx(path: Path) -> tuple[list[dict], str]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl не установлен, .xlsx не поддержан")
        return [], "xlsx"

    rows: list[dict] = []
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return [], "xlsx"
    header: list[str] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:
            header = [str(cell).strip() if cell is not None else f"col{n}"
                      for n, cell in enumerate(row, 1)]
            continue
        d = {header[n]: row[n] for n in range(min(len(header), len(row)))}
        if any((v or "") for v in d.values()):
            rows.append(d)
    wb.close()
    return rows, "xlsx"


def _iter_txt(path: Path) -> tuple[list[dict], str]:
    """Текстовый файл: каждая непустая строка — отдельная запись."""
    rows: list[dict] = []
    with open(path, "r", encoding="utf-8", errors="replace") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append({"text": line})
    return rows, "txt"